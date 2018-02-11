from s4v2 import *

import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
try:
    from openpyxl.cells import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

def save_spreadsheet(filename, data_sample):
    wb = Workbook()
    ws = wb.active
    
    row_index = 1
    for rows in data_sample:
        col_index = 1
        for field in rows:
            col_letter = get_column_letter(col_index)
            ws.cell('{}{}'.format(col_letter, row_index)).value = field
            col_index += 1
        row_index += 1

    wb.save(filename)

kiton_ties = filter_col_by_string(data_from_csv, "brandName", "Kiton")

#save_spreadsheet('_files/kiton_ties.xlsx', kiton_ties)